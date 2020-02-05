import boto3
import json
import xlrd
from openpyxl import workbook, Workbook
from openpyxl import load_workbook
import openpyxl

class ReadReviews:

    loc = ("E://NLP//Reviews.xlsx")
    comprehend = boto3.client(service_name='comprehend', region_name='us-east-1')

    def __init__(self):
        self.wb = xlrd.open_workbook(self.loc)
        self.sheet = self.wb.sheet_by_index(0)
        self.nocols = self.sheet.ncols
        self.lists = [[] for _ in range(self.sheet.ncols)]
        self.sentiments = [[] for _ in range(len(self.lists))]

    def analyze_reviews(self):

        for i in range(self.sheet.ncols):

            for j in range(1, self.sheet.nrows):
                self.lists[i].append(self.sheet.cell_value(j, 0))

        for l in range(len(self.lists)):

            for m in range(len(self.lists[l])):
                self.sentiments[l].append(json.dumps(self.comprehend.detect_sentiment(Text=self.lists[l][m], LanguageCode='en')["Sentiment"]))

    def write_review_in_excel(self):
        xlopen = openpyxl.load_workbook("E://NLP//Reviews.xlsx")

        sheet = xlopen.active
        for i in range(len(self.sentiments)):
            for j in range(len(self.sentiments[i])):
                sheet.cell(row=j + 2, column=self.nocols + i + 1).value = self.sentiments[i][j]
        xlopen.save("E://NLP//Reviews.xlsx")

    def count_reviews_in_excel(self):
        for i in range(len(self.sentiments)):
            positive_sentiment = self.sentiments[i].count('"POSITIVE"')
            mixed_sentiment = self.sentiments[i].count('"MIXED"')
            neutral_sentiment = self.sentiments[i].count('"NEUTRAL"')
            negative_sentiment = self.sentiments[i].count('"NEGATIVE"')


        print(positive_sentiment)

my_reviews = ReadReviews()
my_reviews.analyze_reviews()
my_reviews.write_review_in_excel()
my_reviews.count_reviews_in_excel()




